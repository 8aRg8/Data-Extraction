import openpyxl
import os
os.chdir('@@Folder name where you saved your excel file')
wb = openpyxl.load_workbook('Your excel file name.xlsx')
print(wb.sheetnames)
sheet= wb['@@Your workbook name'] # make sure that each cells contains only text, if you used functions to prepare this workbook, copy it and "past values" in a new workbook 
#print(sheet['B3'].value) #To test that everything works fine

def lists_overlap(a, b): #because we might have multiple papers for each person we define this function
    for i in a:
        if i in b:
            return True
    return False

numbers=[]
names={} #Again I used dictionaries because I had multiple entries for each name, otherwise I could have used lists
titles={}
abstracts={}
for i in range(1,@@): #@@=number of columns
    for j in range(1,@@): #@@=number of rows

        text=str(sheet.cell(row=i, column=j).value)
        if(text):
            if(text.startswith('#')): #the presentation number is followed by #
               cell_list=text.split('::') #with the use of excel functions I seprated names and titles with ::
               number=cell_list[0][1:]           #to remove # from numbers
               number = number.lstrip()
               if cell_list[1] in names:
                   names[cell_list[1]].append(number)
               else:
                   names[cell_list[1]]= [number]
               titles[cell_list[2]]= [number]
               abstracts[cell_list[3]]= [number]
            else:
                continue


#print("here is", names) #To check everything works fine
            #Now I want to print this information nicely
sum=0
for k, v in sorted(names.items()):
    if(len(k)>1 and len(v)>0): #Some cells might have no enteries for abstract or title
        for kk, vv in titles.items():
            if(lists_overlap(v,vv)):
                for s in v:
                    print(s, k,'\n', kk, '\n')
                    v.remove(s)                  #to avoid duplication for multiple enteries per person
                sum=sum+1



num=0


for k, v in abstracts.items():
    if(len(k)>1):
        for s in v:
            print(s, k,'\n')
            num = num+1
        

#print(len(names))
#print(len(abstracts))
print(sum, num) #To make sure we have printed everything
